import openpyxl, json, re

wb = openpyxl.load_workbook(r"C:\Users\000\Downloads\Insurance Comparison.xlsx", data_only=True)
OUTPUT = r"C:\Users\000\.openclaw\workspace\cfp-malaysia\data\products.json"

def clean(v):
    if v is None: return ""
    return str(v).strip()

def parse_age(age_str):
    if not age_str or str(age_str) in ("Not specified", "N/A", "", "Varies"): return None
    s = str(age_str).replace(" years old", "").replace(" years", "").replace(" days", "").strip()
    parts = re.split(r"[-]", s)
    try:
        if len(parts) == 2:
            return {"min": int(float(parts[0].strip())), "max": int(float(parts[1].strip()))}
        return {"min": int(float(parts[0].strip())), "max": int(float(parts[0].strip()))}
    except: return None

def parse_term(term_str):
    if not term_str or str(term_str) in ("Not specified", "N/A", ""): return {"type": "unknown", "value": None}
    s = str(term_str)
    if "lifetime" in s.lower() or "100 years old" in s: return {"type": "lifetime", "value": 100}
    m = re.search(r"age\s*(\d+)", s, re.IGNORECASE)
    if m: return {"type": "age", "value": int(m.group(1))}
    m = re.search(r"(\d+)\s*yrs?", s, re.IGNORECASE)
    if m: return {"type": "years", "value": int(m.group(1))}
    return {"type": "unknown", "value": None}

def parse_rm(s):
    if not s or str(s) in ("Not specified", "N/A", "", "No Limit", "Varies", "Varies by age", "Varies by sust.", "Varies by SA"): return None
    s2 = str(s).replace(",", "").replace(" ", "")
    m = re.search(r"RM?([\d.]+)\s*million", s2, re.IGNORECASE)
    if m: return int(float(m.group(1)) * 1_000_000)
    m = re.search(r"RM?([\d.]+)\s*k", s2, re.IGNORECASE)
    if m: return int(float(m.group(1)) * 1_000)
    m = re.search(r"RM?([\d.]+)", s2, re.IGNORECASE)
    if m: return int(float(m.group(1)))
    return None

catalog = {"meta": {"source": "Insurance Comparison.xlsx", "total_products": 0, "by_category": {}}, "products": []}

def is_header(v):
    if not v: return False
    sv = str(v).strip()
    return sv == "Product Name" or sv.startswith("Product Name ") or "Insurance Provider" in sv

def is_marker(v):
    if not v: return False
    sv = str(v).strip()
    markers = ["Deepsek", "Ringgit", "ringgit", "deepsek", "extracted", "pHS", "PHS"]
    return any(m in sv for m in markers)

def get_header_row_map(ws, header_row, max_col):
    header_map = {}
    for c in range(1, max_col + 1):
        h = clean(ws.cell(header_row, c).value)
        if h:
            header_map[c] = h
    return header_map

def process_sheet(ws, cat_key, id_prefix, extra_fn=None, stop_on_second_header=False):
    max_col = ws.max_column
    max_row = ws.max_row
    current_header_map = {}
    cnt = 0
    products = []
    seen_names = set()
    header_count = 0

    for r in range(1, max_row + 1):
        first_cell_raw = ws.cell(r, 1).value
        if is_header(first_cell_raw):
            header_count += 1
            if stop_on_second_header and header_count > 1:
                break
            current_header_map = get_header_row_map(ws, r, max_col)
            continue
        first_cell = clean(first_cell_raw)
        if not first_cell or first_cell in ("None", "") or is_marker(first_cell):
            continue
        if not current_header_map:
            continue
        row = {}
        for c, key in current_header_map.items():
            row[key] = clean(ws.cell(r, c).value)
        name = row.get("Product Name", row.get("Insurance Product Name", ""))
        if not name: continue
        norm_name = name.strip().lower()
        if norm_name in seen_names: continue
        seen_names.add(norm_name)
        p = {"id": f"{id_prefix}-{cnt+1}", "category": cat_key, "name": name.strip()}
        if extra_fn:
            p = extra_fn(p, row, parse_age, parse_term, parse_rm)
        else:
            p.update({"provider": row.get("Provider", row.get("Insurer", "")), "keyFeatures": row.get("Key Features", "")})
        products.append(p)
        cnt += 1
    return products

# LIFE — stop after first duplicate header (second "Product Name" header at row 39)
ws = wb["Life Comparison"]
def life_extra(p, row, parse_age, parse_term, parse_rm):
    age = parse_age(row.get("Min Entry Age", row.get("Entry Age", "")))
    return {
        **p,
        "provider": row.get("Provider", ""),
        "coverageTerm": row.get("Coverage Term", ""), "coverageTermParsed": parse_term(row.get("Coverage Term", "")),
        "minPremium": row.get("Min. Premium", ""), "minPremiumAmount": parse_rm(row.get("Min. Premium", "")),
        "minSumAssured": row.get("Min. Sum Assured", row.get("Min. SA", "")), "minSumAssuredAmount": parse_rm(row.get("Min. Sum Assured", row.get("Min. SA", ""))),
        "maxPremiumSumAssured": row.get("Max. Premium + Max. SA", ""),
        "parNonPar": row.get("Par / Non-Par", ""), "ilNonIl": row.get("IL / Non-IL", ""),
        "guaranteedBenefit": row.get("Guaranteed Benefit", ""), "nonGuaranteedBenefit": row.get("Non-Guaranteed Benefit", ""),
        "cashValue": row.get("Cash Value", ""), "convertible": row.get("Convertible", ""), "renewable": row.get("Renewable", ""),
        "riders": row.get("Riders", ""), "tpd": row.get("TPD", ""), "suicideClause": row.get("Suicide Clause", ""),
        "entryAgeMin": age["min"] if age else None, "entryAgeMax": age["max"] if age else None,
        "maxAgeMaturity": row.get("Max Age Maturity/Renewal", ""),
        "keyFeatures": row.get("Key Features", "") or row.get("Notes", ""),
    }

life_products = process_sheet(ws, "life", "life", life_extra, stop_on_second_header=True)
catalog["products"].extend(life_products)
catalog["meta"]["by_category"]["life"] = len(life_products)

# CI
ws = wb["CI Comparison"]
def ci_extra(p, row, parse_age, parse_term, parse_rm):
    age = parse_age(row.get("Entry Age", ""))
    return {
        **p,
        "provider": row.get("Provider", ""), "planType": row.get("Plan Type", ""),
        "coverageTerm": row.get("Coverage Term / Max Age", ""), "coverageTermParsed": parse_term(row.get("Coverage Term / Max Age", "")),
        "minSumAssured": row.get("Min / Max Sum Assured", ""), "minSumAssuredAmount": parse_rm(row.get("Min / Max Sum Assured", "")),
        "entryAgeMin": age["min"] if age else None, "entryAgeMax": age["max"] if age else None,
        "noOfIllnesses": row.get("No. of Illnesses", ""), "earlyStageCoverage": row.get("Early Stage Coverage", ""),
        "keyFeatures": row.get("Key Features", ""),
    }

ci_products = process_sheet(ws, "critical_illness", "ci", ci_extra)
catalog["products"].extend(ci_products)
catalog["meta"]["by_category"]["critical_illness"] = len(ci_products)

# MEDICAL
ws = wb["Medical Comparison"]
def med_extra(p, row, parse_age, parse_term, parse_rm):
    return {
        **p,
        "provider": row.get("Provider", ""), "planType": row.get("Plan Type", ""),
        "coverageTerm": row.get("Coverage Term / Max Age", ""), "coverageTermParsed": parse_term(row.get("Coverage Term / Max Age", "")),
        "annualLimit": row.get("Annual Limit", ""), "annualLimitAmount": parse_rm(row.get("Annual Limit", "")),
        "lifetimeLimit": row.get("Lifetime Limit", ""), "lifetimeLimitAmount": parse_rm(row.get("Lifetime Limit", "")),
        "roomAndBoard": row.get("Room & Board", ""), "deductible": row.get("Deductible / Co-Insurance", ""),
        "keyFeatures": row.get("Key Features", ""),
    }

med_products = process_sheet(ws, "medical", "medical", med_extra)
catalog["products"].extend(med_products)
catalog["meta"]["by_category"]["medical"] = len(med_products)

# SAVINGS
ws = wb["Savings Endowment Retirement"]
def sav_extra(p, row, parse_age, parse_term, parse_rm):
    return {
        **p,
        "planType": row.get("Plan Type", ""), "provider": row.get("Insurer", row.get("Provider", "")),
        "premiumTerm": row.get("Premium Term", ""), "coveragePeriod": row.get("Coverage Period", ""),
        "minAnnualPremium": row.get("Min. Annual Premium", ""), "minAnnualPremiumAmount": parse_rm(row.get("Min. Annual Premium", "")),
        "keyFeatures": row.get("Key Features", ""),
    }

sav_products = process_sheet(ws, "savings_endowment_retirement", "savings", sav_extra)
catalog["products"].extend(sav_products)
catalog["meta"]["by_category"]["savings_endowment_retirement"] = len(sav_products)

catalog["meta"]["total_products"] = len(catalog["products"])

with open(OUTPUT, "w", encoding="utf-8") as f:
    json.dump(catalog, f, ensure_ascii=False, indent=2)

print(f"Done! {catalog['meta']['total_products']} products")
for k, v in catalog["meta"]["by_category"].items():
    print(f"  {k}: {v}")
